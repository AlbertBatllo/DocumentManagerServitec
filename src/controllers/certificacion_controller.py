from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple
from models.certificacion_document import CertificacionDocument, CertificacionEntry, CertificacionRepository
from models.licitacion_document import LicitacionDocument, LicitacionRepository
from models.document_summary import CertificacionSummary, create_certificacion_summaries_from_manifest
from utils.certificacion_file_manager import CertificacionFileManager
from utils.trash import move_to_trash
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class CertificacionController:
    """Controller for managing certificacion documents"""
    
    def __init__(self, base_path: str, project_path: Path):
        self.base_path = Path(base_path)
        self.project_path = Path(project_path)
        self.manifest_path = self.base_path / "certificaciones_manifest.json"
        
        # Create base directory if it doesn't exist
        self.base_path.mkdir(parents=True, exist_ok=True)
        
        # Initialize repository
        self.repository = CertificacionRepository(self.manifest_path)
        
        # Reference to licitacion repository for adicionales
        # Try to initialize SQLite licitacion controller first, fallback to JSON
        try:
            from controllers.sqlite_licitacion_controller import SQLiteLicitacionController
            self.licitacion_controller = SQLiteLicitacionController(self.project_path)
            self.licitacion_repo = None  # We'll use the controller instead
        except Exception:
            # Fallback to JSON-based repository
            from utils.path_helper import PathHelper
            pm_path = PathHelper.get_project_manager_path(self.project_path)
            self.licitacion_manifest = pm_path / "presupuestos" / "manifest.json"
            if self.licitacion_manifest.exists():
                self.licitacion_repo = LicitacionRepository(self.licitacion_manifest)
            else:
                self.licitacion_repo = None
            self.licitacion_controller = None
        
        # Initialize file manager
        licitaciones_path = self.project_path / "03_Presupuestos"
        self.file_manager = CertificacionFileManager(self.base_path, licitaciones_path)
        
        # Ensure company folders exist for all existing certificaciones
        self._ensure_company_folders_exist()

    def create_from_licitacion(self, licitacion_name: str, presupuesto_contratado: Optional[float] = None) -> bool:
        """Create a new Certificacion from an approved Licitacion
        
        Args:
            licitacion_name: Name of the licitacion document
            presupuesto_contratado: Optional override for contracted budget. If not provided, uses valor from licitacion
        """
        licitacion = None
        
        # Try SQLite controller first
        if self.licitacion_controller:
            licitacion = self.licitacion_controller.get_document(licitacion_name)
        # Fallback to JSON repository
        elif self.licitacion_repo:
            licitacion = self.licitacion_repo.get_document(licitacion_name)
        else:
            print("Error: No licitacion data source available")
            return False
        
        if not licitacion:
            print(f"Error: Licitacion {licitacion_name} not found")
            return False
        
        # Use the same name as the licitacion for the certificacion
        cert_name = licitacion_name
        if self.repository.document_exists(cert_name):
            print(f"Certificacion for {licitacion_name} already exists")
            return False
        
        # Determine presupuesto_contratado - use valor from licitacion if not overridden
        if presupuesto_contratado is None:
            if hasattr(licitacion, 'valor') and licitacion.valor:
                presupuesto_contratado = licitacion.valor
            else:
                print(f"Error: No valor found in licitacion {licitacion_name} and no presupuesto_contratado provided")
                return False
        
        # Create new Certificacion with the same name as the licitacion
        # Handle both SQLite and JSON licitacion documents
        if hasattr(licitacion, 'lote'):  # SQLite document
            certificacion = CertificacionDocument(
                nombre=cert_name,
                lote=licitacion.lote,
                empresa=licitacion.company,
                presupuesto_contratado=presupuesto_contratado,
                licitacion_name=licitacion_name,
                current_state="S0"
            )
        else:  # JSON document - use the factory method
            certificacion = CertificacionDocument.create_from_licitacion(licitacion, presupuesto_contratado)
            certificacion.nombre = cert_name  # Set the certificacion name to match licitacion
        
        # Create initial entry (version 0)
        initial_entry = CertificacionEntry(
            numero_certificacion=0,
            fecha=datetime.now().strftime("%Y-%m-%d"),
            importe_certificado=0.0,
            retencion=0.0,
            cuenta_prorrata=0.0,
            adicionales_ids=[],
            total_adicionales=0.0,
            total_certificado=0.0,
            porcentaje_completado=0.0,
            author="Sistema",
            notes="Certificación inicial creada automáticamente"
        )
        
        certificacion.add_entry(initial_entry)
        
        # Save to repository
        self.repository.add_document(certificacion)
        
        # Update licitacion with presupuesto_contratado
        if self.licitacion_controller:
            # SQLite controller - presupuesto_contratado should already be set
            pass
        elif self.licitacion_repo:
            # JSON repository
            licitacion.presupuesto_contratado = presupuesto_contratado
            self.licitacion_repo.update_document(licitacion_name, licitacion)
        
        return True

    def create_certificacion_from_adicional(self, cert_name: str, lote: str, company: str,
                                           adicional_name: str, importe: float, notes: str = "") -> str:
        """Create a new Certificacion from an approved adicional. Returns success message."""
        
        # Check if already exists
        if self.repository.document_exists(cert_name):
            raise ValueError(f"Certificacion {cert_name} already exists")
        
        # Create new Certificacion document
        certificacion = CertificacionDocument(
            nombre=cert_name,
            lote=lote,
            empresa=company,
            presupuesto_contratado=importe,  # Use adicional amount as base
            licitacion_name=adicional_name,  # Reference to the adicional
            current_state="S0"
        )
        
        # Create initial entry with the adicional amount
        initial_entry = CertificacionEntry(
            numero_certificacion=0,
            fecha=datetime.now().strftime("%Y-%m-%d"),
            importe_certificado=importe,
            retencion=0.0,
            cuenta_prorrata=0.0,
            adicionales_ids=[adicional_name],
            total_adicionales=importe,
            total_certificado=importe,
            porcentaje_completado=100.0,  # Adicionales are typically 100%
            author="Sistema",
            notes=f"Certificación creada desde adicional: {notes}"
        )
        
        certificacion.add_entry(initial_entry)
        
        # Save to repository
        self.repository.add_document(certificacion)
        
        return f"✓ Certificación {cert_name} creada desde adicional {adicional_name}"

    def add_monthly_certification(self, certificacion_name: str, 
                                importe_certificado: float,
                                retencion: float,
                                cuenta_prorrata: float,
                                adicionales_ids: List[str],
                                author: str,
                                notes: str = "") -> bool:
        """Add a monthly certification entry"""
        certificacion = self.repository.get_document(certificacion_name)
        if not certificacion:
            print(f"Error: Certificacion {certificacion_name} not found")
            return False
        
        # Calculate total adicionales
        total_adicionales = 0.0
        if adicionales_ids:
            for adicional_id in adicionales_ids:
                adicional = None
                # Try SQLite controller first
                if self.licitacion_controller:
                    adicional = self.licitacion_controller.get_document(adicional_id)
                # Fallback to JSON repository
                elif self.licitacion_repo:
                    adicional = self.licitacion_repo.get_document(adicional_id)
                
                if adicional:
                    # Use valor first, fallback to importe_adicional
                    if hasattr(adicional, 'valor') and adicional.valor:
                        total_adicionales += adicional.valor
                    elif hasattr(adicional, 'importe_adicional') and adicional.importe_adicional:
                        total_adicionales += adicional.importe_adicional
        
        # Create new entry
        nuevo_numero = certificacion.numero_certificacion_actual + 1
        nueva_entrada = CertificacionEntry(
            numero_certificacion=nuevo_numero,
            fecha=datetime.now().strftime("%Y-%m-%d"),
            importe_certificado=importe_certificado,
            retencion=retencion,
            cuenta_prorrata=cuenta_prorrata,
            adicionales_ids=adicionales_ids,
            total_adicionales=total_adicionales,
            total_certificado=importe_certificado + total_adicionales,
            porcentaje_completado=0.0,  # Will be calculated in add_entry
            author=author,
            notes=notes
        )
        
        certificacion.add_entry(nueva_entrada)
        self.repository.update_document(certificacion_name, certificacion)
        
        return True

    def get_all_certificaciones(self) -> List[CertificacionDocument]:
        """Get all certificaciones"""
        return self.repository.get_all_documents()
    
    def get_certificacion_summaries(self) -> List[CertificacionSummary]:
        """
        Get lightweight certificacion summaries for fast dashboard loading.
        
        This method provides significant performance improvement by:
        - Reading JSON only once
        - Skipping full CertificacionDocument object creation
        - Extracting only current month/year state information
        - Avoiding complex monthly history processing
        
        Returns:
            List of CertificacionSummary objects with essential data for dashboard display
        """
        try:
            # Read manifest file directly without creating CertificacionDocument objects
            from utils.file_manager import FileManager
            manifest_data = FileManager.safe_json_read(str(self.manifest_path))
            
            # Create summaries efficiently from raw JSON data
            summaries = create_certificacion_summaries_from_manifest(manifest_data)
            
            return summaries
            
        except Exception as e:
            print(f"Error loading certificacion summaries: {e}")
            # Fallback to empty list - allows graceful degradation
            return []

    def get_certificacion(self, certificacion_name: str) -> Optional[CertificacionDocument]:
        """Get a specific certificacion"""
        return self.repository.get_document(certificacion_name)

    def get_available_adicionales(self, licitacion_name: str) -> List[Dict[str, Any]]:
        """Get available adicionales for a licitacion that haven't been used in certifications yet."""
        
        # Get all adicionales related to this licitacion
        adicionales = []
        all_docs = []
        
        # Try SQLite controller first
        if self.licitacion_controller:
            all_docs = self.licitacion_controller.get_all_documents()
            for doc in all_docs:
                if (hasattr(doc, 'parent_licitacion_name') and 
                    doc.parent_licitacion_name == licitacion_name and 
                    hasattr(doc, 'licitacion_document_type') and
                    doc.licitacion_document_type == "adicional"):
                    
                    # Get importe from valor first, then importe_adicional
                    importe = 0.0
                    if hasattr(doc, 'valor') and doc.valor:
                        importe = doc.valor
                    elif hasattr(doc, 'importe_adicional') and doc.importe_adicional:
                        importe = doc.importe_adicional
                    
                    if importe > 0:
                        adicionales.append({
                            'id': doc.name,
                            'name': doc.name,
                            'company': doc.company if hasattr(doc, 'company') else '',
                            'importe_adicional': importe
                        })
        # Fallback to JSON repository
        elif self.licitacion_repo:
            for doc in self.licitacion_repo.get_all_documents():
                if (hasattr(doc, 'parent_licitacion_name') and 
                    doc.parent_licitacion_name == licitacion_name and 
                    doc.document_type == "adicionales" and 
                    hasattr(doc, 'importe_adicional') and
                    doc.importe_adicional is not None):
                    
                    adicionales.append({
                        'id': doc.name,
                        'name': doc.name,
                        'company': doc.company,
                        'importe_adicional': doc.importe_adicional
                    })
        
        # Filter out adicionales already used in certifications
        certificacion = self.repository.get_document(licitacion_name)
        if certificacion:
            used_adicionales = set()
            for entry in certificacion.entries:
                used_adicionales.update(entry.adicionales_ids)
            
            # Return only unused adicionales
            available = [a for a in adicionales if a['id'] not in used_adicionales]
            return available
        
        return adicionales

    def create_monthly_certificacion(self, certificacion_data: Dict[str, Any]) -> str:
        """Create a new monthly certification entry."""
        try:
            certificacion_id = certificacion_data['certificacion_id']
            certificacion = self.repository.get_document(certificacion_id)
            
            if not certificacion:
                raise ValueError(f"Certificación {certificacion_id} no encontrada")
            
            # Get next certification number
            next_number = certificacion.numero_certificacion_actual + 1
            
            # Create the entry
            entry = CertificacionEntry(
                numero_certificacion=next_number,
                fecha=certificacion_data['fecha'],
                importe_certificado=certificacion_data['importe_certificado'],
                retencion=certificacion_data['retencion'],
                cuenta_prorrata=certificacion_data['cuenta_prorrata'],
                adicionales_ids=certificacion_data['adicionales_ids'],
                total_adicionales=certificacion_data['total_adicionales'],
                total_certificado=certificacion_data['importe_certificado'] + certificacion_data['total_adicionales'],
                porcentaje_completado=0.0,  # Will be calculated by add_entry
                author=certificacion_data['author'],
                notes=certificacion_data['notes']
            )
            
            # Add entry to certificacion
            certificacion.add_entry(entry)
            
            # Update repository
            self.repository.update_document(certificacion_id, certificacion)
            
            adicionales_info = ""
            if entry.adicionales_ids:
                adicionales_info = f" (incluye {len(entry.adicionales_ids)} adicionales: {entry.total_adicionales:,.2f} €)"
            
            return f"✓ Certificación #{next_number} creada: {entry.importe_certificado:,.2f} €{adicionales_info}"
            
        except Exception as e:
            raise Exception(f"Error al crear certificación mensual: {str(e)}")

    def export_to_excel(self, output_path: Optional[str] = None) -> str:
        """Export all certificaciones to Excel with proper formatting"""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(self.base_path / f"certificaciones_{timestamp}.xlsx")
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Certificaciones"
        
        # Define headers
        headers = [
            "Nº", "LOTE", "INDUSTRIAL", "PRESUPUESTO",
            "IMPORTE CERTIFICADO", "%COMPLETADO", "RETENCIÓN",
            "CUENTA PRORRATA", "ADICIONAL", "TOTAL CERTIFICADO"
        ]
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Write data
        row_num = 2
        certificaciones = self.get_all_certificaciones()
        
        # Sort by lote number
        certificaciones.sort(key=lambda x: x.lote_number)
        
        for cert in certificaciones:
            latest = cert.latest_entry if cert.latest_entry else None
            
            ws.cell(row=row_num, column=1, value=cert.lote_number).border = border_style
            ws.cell(row=row_num, column=2, value=cert.lote).border = border_style
            ws.cell(row=row_num, column=3, value=cert.empresa).border = border_style
            
            # Financial columns with currency format
            presupuesto_cell = ws.cell(row=row_num, column=4, value=cert.presupuesto_contratado)
            presupuesto_cell.number_format = '#,##0.00 €'
            presupuesto_cell.border = border_style
            
            importe_cell = ws.cell(row=row_num, column=5, value=cert.cumulative_certificado)
            importe_cell.number_format = '#,##0.00 €'
            importe_cell.border = border_style
            
            # Percentage with percentage format
            porcentaje_cell = ws.cell(row=row_num, column=6, value=cert.porcentaje_completado_actual / 100)
            porcentaje_cell.number_format = '0.00%'
            porcentaje_cell.border = border_style
            
            # Retention and prorrata
            retencion = latest.retencion if latest else 0
            retencion_cell = ws.cell(row=row_num, column=7, value=retencion)
            retencion_cell.number_format = '#,##0.00 €'
            retencion_cell.border = border_style
            
            prorrata = latest.cuenta_prorrata if latest else 0
            prorrata_cell = ws.cell(row=row_num, column=8, value=prorrata)
            prorrata_cell.number_format = '#,##0.00 €'
            prorrata_cell.border = border_style
            
            # Adicionales
            adicionales_cell = ws.cell(row=row_num, column=9, value=cert.cumulative_adicionales)
            adicionales_cell.number_format = '#,##0.00 €'
            adicionales_cell.border = border_style
            
            # Total
            total_cell = ws.cell(row=row_num, column=10, value=cert.total_certificado_global)
            total_cell.number_format = '#,##0.00 €'
            total_cell.border = border_style
            
            row_num += 1
        
        # Add totals row
        if certificaciones:
            ws.cell(row=row_num, column=2, value="TOTALES").font = Font(bold=True)
            
            # Calculate totals
            total_presupuesto = sum(c.presupuesto_contratado for c in certificaciones)
            total_importe = sum(c.cumulative_certificado for c in certificaciones)
            total_retencion = sum((c.latest_entry.retencion if c.latest_entry else 0) for c in certificaciones)
            total_prorrata = sum((c.latest_entry.cuenta_prorrata if c.latest_entry else 0) for c in certificaciones)
            total_adicionales = sum(c.cumulative_adicionales for c in certificaciones)
            total_global = sum(c.total_certificado_global for c in certificaciones)
            
            # Write totals with formatting
            for col, value in [
                (4, total_presupuesto),
                (5, total_importe),
                (7, total_retencion),
                (8, total_prorrata),
                (9, total_adicionales),
                (10, total_global)
            ]:
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.number_format = '#,##0.00 €'
                cell.font = Font(bold=True)
                cell.border = border_style
            
            # Average percentage
            if total_presupuesto > 0:
                avg_percentage = (total_importe / total_presupuesto)
                percentage_cell = ws.cell(row=row_num, column=6, value=avg_percentage)
                percentage_cell.number_format = '0.00%'
                percentage_cell.font = Font(bold=True)
                percentage_cell.border = border_style
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, ValueError, TypeError) as e:
                    print(f"Warning: Could not get cell value length: {e}")
                    pass
            
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_path)
        
        return output_path

    def get_financial_summary(self) -> Dict[str, Any]:
        """Get financial summary across all certificaciones"""
        return self.repository.get_financial_summary()

    def link_adicional_to_licitacion(self, adicional_id: str, parent_licitacion_name: str, 
                                    importe: float) -> bool:
        """Link an adicional to its parent licitacion"""
        if not self.licitacion_repo:
            return False
        
        adicional = self.licitacion_repo.get_document(adicional_id)
        if not adicional:
            return False
        
        adicional.parent_licitacion_name = parent_licitacion_name
        adicional.importe_adicional = importe
        
        self.licitacion_repo.update_document(adicional_id, adicional)
        return True
    
    def create_monthly_certificacion_with_files(self, certificacion_data: Dict[str, Any]) -> str:
        """
        Create a new monthly certificacion with file management
        
        Args:
            certificacion_data: Dictionary containing:
                - certificacion_id: ID of the base certificacion
                - fecha: Date string
                - importe_certificado: Amount certified
                - retencion: Retention amount
                - cuenta_prorrata: Prorate account amount
                - adicionales_ids: List of adicional IDs
                - total_adicionales: Total adicionales amount
                - notes: Notes
                - author: Author name
                - attached_files: List of file paths to attach
                
        Returns:
            Success message string
        """
        try:
            # Get the base certificacion
            certificacion = self.repository.get_document(certificacion_data['certificacion_id'])
            if not certificacion:
                raise ValueError(f"Certificacion {certificacion_data['certificacion_id']} not found")
            
            # Ensure company folder exists
            company_folder = self.file_manager.create_company_folder(certificacion)
            
            # Calculate next certificacion number
            next_numero = len(certificacion.entries)
            
            # Generate version based on current date or provided date
            fecha_str = certificacion_data['fecha']
            try:
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
                version = fecha_obj.strftime('%m_%Y')  # Format: MM_YYYY
            except Exception as e:
                from utils.error_logger import logger
                logger.warning(f"Failed to parse date for certificacion version", {"fecha_str": fecha_str, "error": str(e)})
                version = datetime.now().strftime('%m_%Y')
            
            # Handle file attachments
            file_results = []
            if certificacion_data.get('attached_files'):
                print(f"Attaching {len(certificacion_data['attached_files'])} files...")
                file_results = self.file_manager.attach_files_to_certificacion(
                    certificacion, 
                    next_numero, 
                    version,
                    certificacion_data['attached_files']
                )
                
                # Check for file errors
                failed_files = [r for r in file_results if r['status'] == 'error']
                if failed_files:
                    error_msg = "Errores al adjuntar archivos:\n"
                    for result in failed_files:
                        error_msg += f"• {result['source']}: {result['message']}\n"
                    raise Exception(error_msg)
            
            # Handle adicionales file movement
            adicionales_file_results = []
            if certificacion_data.get('adicionales_ids'):
                print(f"Moving files for {len(certificacion_data['adicionales_ids'])} adicionales...")
                adicionales_file_results = self.file_manager.move_adicionales_files(
                    certificacion,
                    certificacion_data['adicionales_ids']
                )
            
            # Create the certificacion entry
            entry_result = self.create_monthly_certificacion(certificacion_data)
            
            # Prepare success message with file operation results
            success_msg = f"Certificación mensual creada exitosamente.\n"
            
            if file_results:
                success_files = [r for r in file_results if r['status'] == 'success']
                if success_files:
                    success_msg += f"\nArchivos adjuntos: {len(success_files)}\n"
                    for result in success_files:
                        success_msg += f"• {result['filename']}\n"
                
                warning_files = [r for r in file_results if r['status'] == 'warning']
                if warning_files:
                    success_msg += f"\nAdvertencias de archivos:\n"
                    for result in warning_files:
                        success_msg += f"• {result['message']}\n"
            
            if adicionales_file_results:
                moved_files = [r for r in adicionales_file_results if r['status'] == 'success']
                if moved_files:
                    success_msg += f"\nArchivos de adicionales movidos: {len(moved_files)}\n"
            
            # Add folder information
            folder_info = self.file_manager.get_folder_size_info(certificacion)
            success_msg += f"\nCarpeta: {Path(folder_info['folder_path']).name}"
            success_msg += f"\nTotal archivos: {folder_info['file_count']}"
            success_msg += f"\nTamaño total: {folder_info['total_size_mb']:.1f} MB"
            
            return success_msg
            
        except Exception as e:
            # Rollback: try to clean up any files that were copied
            try:
                if 'file_results' in locals():
                    for result in file_results:
                        if result['status'] == 'success' and 'destination' in result:
                            dest_path = Path(result['destination'])
                            if dest_path.exists():
                                dest_path.unlink()
            except Exception as rollback_error:
                # Log rollback errors - they might indicate file system issues
                from utils.error_logger import logger
                logger.warning(f"File cleanup failed during error recovery", {"error": str(rollback_error)})
            
            raise Exception(f"Error creating monthly certificacion: {str(e)}")
    
    def get_adicionales_files(self, adicionales_ids: List[str]) -> Dict[str, List[str]]:
        """
        Get file paths for given adicionales IDs
        
        Args:
            adicionales_ids: List of adicional licitacion IDs
            
        Returns:
            Dictionary mapping adicional ID to list of file paths
        """
        result = {}
        
        for adicional_id in adicionales_ids:
            if not self.licitacion_repo:
                result[adicional_id] = []
                continue
                
            # Get adicional document to find parent licitacion
            adicional = self.licitacion_repo.get_document(adicional_id)
            if adicional and adicional.parent_licitacion_name:
                files = self.file_manager.get_available_adicionales_files(adicional.parent_licitacion_name)
                result[adicional_id] = [str(f) for f in files]
            else:
                result[adicional_id] = []
        
        return result
    
    def list_certificacion_files(self, certificacion_id: str) -> Dict[str, List[str]]:
        """
        List all files associated with a certificacion
        
        Args:
            certificacion_id: ID of the certificacion
            
        Returns:
            Dictionary with 'certificacion' and 'adicionales' file lists
        """
        certificacion = self.repository.get_document(certificacion_id)
        if not certificacion:
            return {'certificacion': [], 'adicionales': []}
        
        files_info = self.file_manager.list_certificacion_files(certificacion)
        
        return {
            'certificacion': [str(f) for f in files_info['certificacion']],
            'adicionales': [str(f) for f in files_info['adicionales']]
        }
    
    def _ensure_company_folders_exist(self) -> None:
        """Ensure company folders exist for all existing certificaciones"""
        try:
            certificaciones = self.get_all_certificaciones()
            for certificacion in certificaciones:
                self.file_manager.get_company_folder_path(certificacion)
        except Exception as e:
            print(f"Warning: Could not ensure company folders exist: {e}")
    
    def cleanup_empty_folders(self) -> List[str]:
        """Clean up empty company folders and return list of removed folders"""
        return self.file_manager.cleanup_empty_folders()
    
    def get_company_folder_info(self, certificacion_id: str) -> Optional[Dict[str, Any]]:
        """Get information about the company folder for a certificacion"""
        certificacion = self.repository.get_document(certificacion_id)
        if not certificacion:
            return None
        
        return self.file_manager.get_folder_size_info(certificacion)
    
    def update_certificacion_state(self, certificacion_id: str, new_state: str, author: str = "Usuario") -> str:
        """Update the state of a certificacion"""
        from models.certificacion_document import CERTIFICACION_STATES
        
        # Validate state
        if new_state not in CERTIFICACION_STATES:
            raise ValueError(f"Estado '{new_state}' no válido. Debe ser uno de: {CERTIFICACION_STATES}")
        
        # Get certificacion
        certificacion = self.repository.get_document(certificacion_id)
        if not certificacion:
            raise ValueError(f"Certificación '{certificacion_id}' no encontrada")
        
        # Store old state for logging
        old_state = certificacion.current_state
        
        # Update state
        certificacion.update_state(new_state)
        
        # Save changes
        self.repository.update_document(certificacion_id, certificacion)
        
        # Return success message with state names
        old_display = certificacion.get_state_display_name(old_state)
        new_display = certificacion.get_state_display_name(new_state)
        
        return f"Estado actualizado de '{old_display}' a '{new_display}'"
    
    def get_certificacion_state_summary(self) -> Dict[str, int]:
        """Get summary of certificacion counts per state"""
        from models.certificacion_document import CERTIFICACION_STATES
        
        summary = {state: 0 for state in CERTIFICACION_STATES}
        
        for certificacion in self.repository.get_all_documents():
            if certificacion.current_state in summary:
                summary[certificacion.current_state] += 1
        
        return summary
    
    def get_certificaciones_by_state(self, state: str) -> List[CertificacionDocument]:
        """Get all certificaciones in a specific state"""
        return [cert for cert in self.repository.get_all_documents() if cert.current_state == state]
    
    def create_certificacion_from_adicional(self, cert_name: str, lote: str, company: str, 
                                          adicional_name: str, importe: float, 
                                          notes: str = "") -> str:
        """Create a new certificacion from an approved adicional. Returns success message."""
        from models.certificacion_document import CertificacionDocument, CertificacionEntry
        from datetime import datetime
        
        # Check if certificacion already exists
        if self.repository.document_exists(cert_name):
            raise ValueError(f"Ya existe una certificación con el nombre '{cert_name}'")
        
        # Create new certificacion document
        certificacion = CertificacionDocument(
            nombre=cert_name,
            lote=lote,
            empresa=company,
            presupuesto_contratado=0.0,  # Will be updated when base contract is available
            licitacion_name=adicional_name,  # Reference to source adicional
            entries=[]
        )
        
        # Add initial entry with adicional information
        entry = CertificacionEntry(
            numero_certificacion=0,  # Initial certification
            fecha=datetime.now().strftime("%Y-%m-%d"),
            importe_certificado=0.0,  # No base certification amount yet
            retencion=0.0,
            cuenta_prorrata=0.0,
            adicionales_ids=[adicional_name],  # Track source adicional
            total_adicionales=importe,
            total_certificado=importe,  # Only adicional amount for now
            porcentaje_completado=0.0,  # To be determined later
            author="Sistema",
            notes=f"Certificación creada automáticamente desde adicional aprobado: {adicional_name}. {notes}".strip()
        )
        
        certificacion.entries.append(entry)
        
        # Ensure company folder exists
        try:
            self.file_manager.get_company_folder_path(certificacion)
        except Exception as e:
            print(f"Warning: Could not create company folder: {e}")
        
        # Save to repository
        self.repository.add_document(certificacion)
        
        return f"✓ Certificación '{cert_name}' creada desde adicional '{adicional_name}' con importe {importe:,.2f} €"
    
    def _resolve_file_path(self, certificacion_name: str) -> Optional[Path]:
        """Resolve file path for certificacion, preferring stored paths over pattern matching."""
        certificacion = self.repository.get_document(certificacion_name)
        if not certificacion:
            return None
        
        # First priority: Check if we have a stored file_path in the latest entry
        latest_entry = certificacion.latest_entry
        if latest_entry and hasattr(latest_entry, 'file_path') and latest_entry.file_path:
            # Try absolute path first
            stored_path = Path(latest_entry.file_path)
            if stored_path.exists():
                return stored_path
                
            # Try relative to company folder
            try:
                company_folder = self.file_manager.get_company_folder_path(certificacion)
                relative_path = company_folder / latest_entry.file_path
                if relative_path.exists():
                    return relative_path
            except Exception:
                pass
                
            # Try relative to base path
            relative_path = self.base_path / latest_entry.file_path
            if relative_path.exists():
                return relative_path
        
        # Fallback: Use pattern matching for legacy documents without stored paths
        # Determine target folder (usually company folder)
        try:
            company_folder = self.file_manager.get_company_folder_path(certificacion)
        except Exception:
            company_folder = self.base_path
        
        # Determine latest entry number (numero_certificacion) if available
        numero = 0
        if certificacion.entries:
            try:
                numero = max(e.numero_certificacion for e in certificacion.entries)
            except Exception:
                numero = 0
        
        from utils.certificacion_file_manager import CertificacionFileManager
        candidates: List[Path] = []
        # Build primary expected filename
        try:
            primary_name = CertificacionFileManager.generate_certificacion_filename(
                certificacion_id=certificacion.nombre,
                numero_cert=numero,
                date_str=None,
                state=None,
                file_ext="pdf"
            )
            candidates.append(company_folder / primary_name)
        except Exception:
            pass
        
        # Try common extensions and looser patterns
        for ext in ['pdf', 'xlsx', 'xls', 'docx', 'doc']:
            # Simple pattern: startswith certificacion name
            try:
                for fp in company_folder.iterdir():
                    if fp.is_file() and fp.name.startswith(str(certificacion.nombre)) and fp.suffix.lstrip('.') == ext:
                        return fp
            except Exception:
                pass
        
        # Fallback to base path
        for ext in ['pdf', 'xlsx', 'xls', 'docx', 'doc']:
            try:
                for fp in self.base_path.iterdir():
                    if fp.is_file() and fp.name.startswith(str(certificacion.nombre)) and fp.suffix.lstrip('.') == ext:
                        return fp
            except Exception:
                pass
        
        # Finally check explicit candidates
        for c in candidates:
            if c.exists():
                return c
        
        return None

    def open_document_location(self, certificacion_name: str) -> None:
        """Open the file location for a certificacion document."""
        file_path = self._resolve_file_path(certificacion_name)
        if not file_path:
            # Check if this is an empty template certificacion (only has automatic version 0)
            certificacion = self.repository.get_document(certificacion_name)
            if certificacion and certificacion.entries and len(certificacion.entries) == 1:
                entry = certificacion.entries[0]
                if entry.numero_certificacion == 0 and entry.author == "Sistema":
                    raise FileNotFoundError(
                        f"La certificación '{certificacion_name}' es una plantilla automática sin documentos asociados.\n\n"
                        "Para abrir archivos, primero debe:\n"
                        "1. Crear certificaciones mensuales con documentos\n"
                        "2. O subir archivos manualmente a la carpeta de la empresa"
                    )
            
            raise FileNotFoundError(f"No se encontró el archivo para la certificación {certificacion_name}")
        
        self.file_manager.open_file_location(file_path)
    
    def delete_document_entries(self, entry_specs: List[Dict[str, str]]) -> str:
        """Delete specific version/state entries. Returns success message."""
        if not entry_specs:
            raise ValueError("No se proporcionaron entradas para eliminar")
        
        deleted_count = 0
        errors = []
        documents_to_update = {}  # Track which documents need updating
        
        for entry_spec in entry_specs:
            try:
                doc_id = entry_spec['doc_id']
                version = entry_spec['version']
                state = entry_spec['state']
                filename = entry_spec['filename']
                
                # Get the certificacion document (using nombre as key)
                certificacion = self.repository.get_document(doc_id)
                if not certificacion:
                    errors.append(f"Certificación {doc_id} no encontrada")
                    continue
                
                # Move physical file to trash (recoverable)
                try:
                    company_folder = self.file_manager.get_company_folder_path(certificacion)
                    file_path = company_folder / filename
                    if file_path.exists():
                        move_to_trash(file_path, self.project_path)
                    else:
                        # Try alternative paths in case file is in base path
                        alt_path = self.base_path / filename
                        if alt_path.exists():
                            move_to_trash(alt_path, self.project_path)
                except OSError as e:
                    errors.append(f"Error al mover archivo {filename} a la papelera: {e}")
                    continue
                
                # Find and remove the specific entry from the certificacion
                # For certificaciones, we'll match by numero_certificacion (version) and state
                entry_removed = False
                for i, entry in enumerate(certificacion.entries):
                    # Convert version to int for comparison with numero_certificacion
                    try:
                        version_num = int(version)
                        if entry.numero_certificacion == version_num:
                            certificacion.entries.pop(i)
                            entry_removed = True
                            deleted_count += 1
                            documents_to_update[doc_id] = certificacion
                            break
                    except ValueError:
                        # If version is not a number, try string comparison
                        if str(entry.numero_certificacion) == version:
                            certificacion.entries.pop(i)
                            entry_removed = True
                            deleted_count += 1
                            documents_to_update[doc_id] = certificacion
                            break
                
                if not entry_removed:
                    errors.append(f"Entrada certificación #{version} no encontrada para certificación {doc_id}")
                
            except Exception as e:
                errors.append(f"Error al procesar entrada {doc_id} certificación #{version}: {e}")
        
        # Update or remove documents as needed (and clean up SQLite if present)
        for doc_id, certificacion in documents_to_update.items():
            try:
                if not certificacion.entries:
                    # Remove in-memory/JSON record
                    if self.repository.document_exists(doc_id):
                        del self.repository.documents[doc_id]
                    # Best-effort: also remove from SQLite if it exists there
                    try:
                        from utils.project_database_manager import ensure_project_database
                        dbm = ensure_project_database(self.project_path)
                        sqlite_id = dbm.get_document_id("certificaciones", doc_id)
                        if sqlite_id is not None:
                            dbm.delete_document(sqlite_id)
                    except Exception:
                        # Ignore SQLite cleanup errors to avoid blocking UI
                        pass
                else:
                    # Update the document with remaining entries
                    self.repository.update_document(doc_id, certificacion)
            except Exception as e:
                errors.append(f"Error al actualizar certificación {doc_id}: {e}")
        
        # Save repository changes
        if deleted_count > 0:
            self.repository.save()
        
        # Prepare result message
        result_parts = []
        if deleted_count > 0:
            result_parts.append(f"Se eliminaron {deleted_count} certificación(es) exitosamente")
        
        if errors:
            result_parts.append(f"Errores encontrados:\n" + "\n".join(f"• {error}" for error in errors))
        
        if not result_parts:
            return "No se eliminó ninguna certificación"
        
        return "\n\n".join(result_parts)

    def update_document_info(self, old_name: str, new_name: str, new_display_name: str, 
                           new_version: str, new_state: str, author: str, notes: str,
                           autor: str = "", rev_tecnica: str = "", rev_gerencia: str = "") -> str:
        """Update certificacion document general information. Returns success message.
        
        Note: old_name/new_name are actually document names (the primary identifier),
        new_display_name is for display purposes but stored in the name field.
        This maintains compatibility with the correction form that passes old_id/new_id.
        """
        certificacion = self.repository.get_document(old_name)
        if not certificacion:
            raise ValueError(f"No se encontró la certificación con nombre {old_name}")
        
        # Validate new state
        from models.certificacion_document import CERTIFICACION_STATES
        if new_state not in CERTIFICACION_STATES:
            raise ValueError(f"Estado '{new_state}' no válido. Debe ser uno de: {CERTIFICACION_STATES}")
        
        # For certificaciones, we don't typically rename files since they're managed differently
        # But we do update the document information and add a new entry
        
        # If name changed, we need to delete old and create new document record
        if old_name != new_name:
            # Remove old document from repository
            del self.repository.documents[old_name]
            
            # Update document info
            certificacion.nombre = new_name  # Use new_name as both key and internal name
            
            # Create a correction entry for certificaciones
            # For certificaciones, we'll create a new entry with the correction information
            from models.certificacion_document import CertificacionEntry
            from datetime import datetime
            
            # Get next certification number
            next_numero = certificacion.numero_certificacion_actual + 1
            
            correction_entry = CertificacionEntry(
                numero_certificacion=next_numero,
                fecha=datetime.now().strftime("%Y-%m-%d"),
                importe_certificado=0.0,  # This would need to be provided in a real correction
                retencion=0.0,
                cuenta_prorrata=0.0,
                adicionales_ids=[],
                total_adicionales=0.0,
                total_certificado=0.0,
                porcentaje_completado=0.0,
                author=author,
                notes=f"Corrección de información: {notes}"
            )
            
            certificacion.add_entry(correction_entry)
            
            # Add document with new name as key
            self.repository.documents[new_name] = certificacion
            self.repository.save()
        else:
            # Just update existing document (name stays the same)
            certificacion.nombre = new_name  # Use new_name as both key and internal name
            
            # Update state if different
            if hasattr(certificacion, 'current_state') and certificacion.current_state != new_state:
                certificacion.update_state(new_state)
            
            # For existing documents, we can add a note about the correction
            # without creating a new certification entry
            self.repository.update_document(old_name, certificacion)
        
        self.repository.save()
        return "Información de certificación actualizada correctamente"
    
    def get_document_file_path(self, doc_name: str) -> Optional[Path]:
        """Get the current file path for a certificacion document using stored file paths"""
        document = self.repository.get_document(doc_name)
        
        # First, try to get file paths from the document's file_paths field (if it exists)
        if document and hasattr(document, 'file_paths') and document.file_paths:
            import json
            try:
                if isinstance(document.file_paths, str):
                    file_paths = json.loads(document.file_paths)
                else:
                    file_paths = document.file_paths
                
                # Try each stored file path
                for file_path_str in file_paths:
                    if file_path_str:  # Skip empty strings
                        file_path = Path(file_path_str)
                        if file_path.exists():
                            return file_path
            except (json.JSONDecodeError, TypeError):
                pass
        
        # Fallback: try to find files in the certificacion company folder
        if document:
            try:
                company_folder = self.file_manager.get_company_folder_path(document)
                if company_folder.exists():
                    # Look for files in the company folder
                    import glob
                    from utils.file_manager import FileManager
                    sanitized_name = FileManager.sanitize_for_filename(doc_name)
                    
                    # Search patterns to try
                    search_patterns = [
                        f"*{sanitized_name}*",  # Direct name match
                        f"*{doc_name}*",        # Original name match
                        f"*{sanitized_name.replace('_', '*')}*",  # More flexible matching
                    ]
                    
                    for pattern in search_patterns:
                        file_pattern = str(company_folder / pattern)
                        found_files = glob.glob(file_pattern)
                        if found_files:
                            # Return the first matching file
                            return Path(found_files[0])
            except Exception:
                pass
        
        # Final fallback: try to find files in the base certificaciones directory
        import glob
        from utils.file_manager import FileManager
        sanitized_name = FileManager.sanitize_for_filename(doc_name)
        
        # Search patterns to try
        search_patterns = [
            f"*{sanitized_name}*",  # Direct name match
            f"*{doc_name}*",        # Original name match
            f"*{sanitized_name.replace('_', '*')}*",  # More flexible matching
        ]
        
        for pattern in search_patterns:
            file_pattern = str(self.base_path / pattern)
            found_files = glob.glob(file_pattern)
            if found_files:
                # Return the first matching file
                return Path(found_files[0])
        
        return None

    def open_document_location(self, doc_name: str) -> None:
        """Open the folder containing the certificacion files"""
        # First try to get the specific file path
        file_path = self.get_document_file_path(doc_name)
        
        if file_path:
            # Found specific file - use the file manager to open and select it
            self.file_manager.open_file_location(file_path)
        else:
            # No specific file found - try to open the company folder
            document = self.repository.get_document(doc_name)
            if document:
                try:
                    company_folder = self.file_manager.get_company_folder_path(document)
                    if company_folder.exists():
                        # Open the company folder
                        import subprocess
                        import platform
                        
                        folder_path = str(company_folder)
                        
                        try:
                            if platform.system() == "Darwin":  # macOS
                                subprocess.Popen(["open", folder_path])
                            elif platform.system() == "Windows":
                                subprocess.Popen(["explorer", folder_path])
                            else:  # Unsupported platform
                                subprocess.Popen(["explorer", folder_path])
                        except Exception as e:
                            raise RuntimeError(f"No se pudo abrir la ubicación del documento: {e}")
                        return
                except Exception as e:
                    # If company folder access fails, continue to final fallback
                    pass
            
            # Final fallback - open the base certificaciones directory
            import subprocess
            import platform
            
            folder_path = str(self.base_path)
            
            try:
                if platform.system() == "Darwin":  # macOS
                    subprocess.Popen(["open", folder_path])
                elif platform.system() == "Windows":
                    subprocess.Popen(["explorer", folder_path])
                else:  # Unsupported platform
                    subprocess.Popen(["explorer", folder_path])
            except Exception as e:
                raise RuntimeError(f"No se pudo abrir la ubicación del documento: {e}")